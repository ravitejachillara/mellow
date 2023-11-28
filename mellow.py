import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import openpyxl
import re
import dns.resolver

# Email Verification Functions
def is_valid_email_syntax(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None

def has_mx_record(email):
    domain = email.split('@')[-1]
    try:
        dns.resolver.resolve(domain, 'MX')
        return True
    except Exception:
        return False

def verify_email(email):
    if not is_valid_email_syntax(email):
        return "Invalid Syntax"
    if not has_mx_record(email):
        return "Invalid Domain/MX Record"
    return "Valid Email"

# Function to Process Excel File
def process_excel_file(input_file_path, output_file_path, progress_callback):
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active
    total_rows = sheet.max_row - 1

    for index, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=2), start=1):
        email_cell = row[0]
        status_cell = row[1]
        email = email_cell.value

        if email:
            status_cell.value = verify_email(email)

        progress = (index / total_rows) * 100
        progress_callback(progress)

    workbook.save(output_file_path)

# Tkinter UI Class
class EmailVerifierApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mellow - Email Verifier by Ravi Teja Chillara")
        self.geometry("380x140")

        # UI Elements
        ttk.Button(self, text="Open File", command=self.open_file).pack(pady=10)
        ttk.Button(self, text="Save As", command=self.save_file).pack(pady=10)
        self.progress = ttk.Progressbar(self, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress.pack(pady=10)

        # File Paths
        self.input_file_path = ""
        self.output_file_path = ""

    def open_file(self):
        self.input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    def save_file(self):
        self.output_file_path = filedialog.asksaveasfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.input_file_path and self.output_file_path:
            self.start_processing()

    def start_processing(self):
        threading.Thread(target=self.process_file, daemon=True).start()

    def process_file(self):
        process_excel_file(self.input_file_path, self.output_file_path, self.update_progress)
        messagebox.showinfo("Completion", "File processing complete.")

    def update_progress(self, value):
        self.progress['value'] = value
        self.update_idletasks()

if __name__ == "__main__":
    app = EmailVerifierApp()
    app.mainloop()
